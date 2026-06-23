#!/usr/bin/env python3
"""
Generate readable rater materials from the raw JIR/JOR predicates:
  - items_to_rate.csv        : item_id, subject, mr_name, JIR, JOR  (READABLE math)
  - items_raw.csv            : the original raw Java predicates (traceability)
  - rating_sheet_TEMPLATE.xlsx : readable items + a category DROPDOWN (a-j/orphan)
  - items_to_rate.tex        : LaTeX source -> compile to items_to_rate.pdf

Faithful structural rewrite (no semantic guessing):
  ((double) X)            -> X            (drop casts)
  (k * X) / (X * k)       -> k*X
  Math.abs(A - B) < 1e-4  -> A ~= B       (approx-equal within tolerance)
  ==,!=,<=,>=,<,>         -> =,!=,<=,>=,<,>
  && , ||                 -> AND , OR
  i_this_real_f -> this.re1 ; i_other_imag_s -> other.im2
  o_return_real_f -> out.re1 ; o_return_f -> out1 ; i_x_f -> x1 ; etc.
Anything that does not match falls back to the cleaned-but-literal form, so the
rewrite never invents meaning. Run with --verify to print raw vs readable for all.
"""
import csv, re, json, sys, subprocess, pathlib

HERE = pathlib.Path(__file__).parent
SRC = HERE.parent.parent / "supplementary" / "S3_case_study" / "lrca_llm_labels.json"
FAMILIES = list("abcdefghij")

# Authoritative program context (signature + behaviour), so a rater can read the
# variables and judge the relation. Sources: paper SUT (MathSignalClass /
# ComplexSignal) + verified semantics (gcd/lcm/powerSig take |input|; exactLog2(2n)
# = exactLog2(n)+1; hypotSig = sqrt(a^2+b^2)).
PROGRAMS = {
    "ComplexSignal.add.0": "add(other): complex addition; out = this + other (real and imaginary parts added separately). Operands this=(this.re,this.im), other=(other.re,other.im); result out=(out.re,out.im).",
    "MathSignalClass.clamp.0": "clamp(lo, hi, x): clamp x into the interval [lo, hi]; returns min(max(x, lo), hi).",
    "MathSignalClass.exactLog2.0": "exactLog2(n): base-2 logarithm of a positive integer n (so exactLog2(2*n) = exactLog2(n) + 1).",
    "MathSignalClass.gcdSig.0": "gcdSig(a, b): greatest common divisor of |a| and |b| (inputs sign-normalised, then Euclidean gcd).",
    "MathSignalClass.hypotSig.0": "hypotSig(a, b): Euclidean hypotenuse, sqrt(a^2 + b^2).",
    "MathSignalClass.isSequence.0": "isSequence(a, b, c): returns true iff a, b, c form an arithmetic sequence (b - a == c - b).",
    "MathSignalClass.lcmSig.0": "lcmSig(a, b): least common multiple of |a| and |b| (inputs sign-normalised).",
    "MathSignalClass.midpoint.0": "midpoint(a, b): arithmetic midpoint, (a + b) / 2.",
    "MathSignalClass.powerSig.0": "powerSig(base, n): |base| raised to the power n (base sign-normalised before exponentiation).",
    "MathSignalClass.signum.0": "signum(x): sign of x; returns -1, 0, or +1 for x<0, x=0, x>0.",
}

SUB = {"f": "1", "s": "2"}

def rename_vars(s):
    # complex parts first
    s = re.sub(r'i_(this|other)_(real|imag)_([fs])',
               lambda m: "%s.%s%s" % (m.group(1), {"real": "re", "imag": "im"}[m.group(2)], SUB[m.group(3)]), s)
    s = re.sub(r'o_return_(real|imag)_([fs])',
               lambda m: "out.%s%s" % ({"real": "re", "imag": "im"}[m.group(1)], SUB[m.group(2)]), s)
    s = re.sub(r'o_return_([fs])', lambda m: "out%s" % SUB[m.group(1)], s)
    s = re.sub(r'o_(\w+?)_([fs])', lambda m: "%s%s" % (m.group(1), SUB[m.group(2)]), s)
    s = re.sub(r'i_(\w+?)_([fs])', lambda m: "%s%s" % (m.group(1), SUB[m.group(2)]), s)
    return s

def clean(s):
    s = re.sub(r'\(\((?:double|int|long|float)\)\s*([A-Za-z0-9_]+)\)', r'\1', s)  # ((double) X)->X
    s = re.sub(r'\((?:double|int|long|float)\)\s*', '', s)                          # stray casts
    s = re.sub(r'\(\s*([0-9.]+)\s*\*\s*([A-Za-z0-9_]+)\s*\)', r'\1*\2', s)          # (k*X)->k*X
    s = re.sub(r'\(\s*([A-Za-z0-9_]+)\s*\*\s*([0-9.]+)\s*\)', r'\2*\1', s)          # (X*k)->k*X
    return s

def _match_balanced(s, start):
    """start points at '(' ; return index just after the matching ')'."""
    depth = 0
    for i in range(start, len(s)):
        if s[i] == '(':
            depth += 1
        elif s[i] == ')':
            depth -= 1
            if depth == 0:
                return i + 1
    return -1

def rewrite_abs(s):
    out, i = [], 0
    while True:
        m = re.search(r'Math\.abs\(', s[i:])
        if not m:
            out.append(s[i:]); break
        j = i + m.start()
        out.append(s[i:j])
        op = j + len("Math.abs")           # at '('
        close = _match_balanced(s, op)
        if close < 0:
            out.append(s[j:j + 8]); i = j + 8; continue
        inner = s[op + 1:close - 1]
        tail = s[close:]
        mt = re.match(r'\s*<\s*[0-9.]+E?-?[0-9]*', tail)  # the < tol part
        if mt:
            parts = inner.split(' - ')
            atom = ("%s ~= %s" % (parts[0].strip(), parts[1].strip())) if len(parts) == 2 \
                   else ("|%s| < eps" % inner.strip())
            out.append("(" + atom + ")")
            i = close + mt.end()
        else:
            out.append("|" + inner + "|")
            i = close
    return "".join(out)

def to_readable(expr):
    s = clean(expr)
    s = rewrite_abs(s)
    s = s.replace("==", "=").replace("!=", " != ").replace("<=", " <= ").replace(">=", " >= ")
    s = s.replace("&&", " AND ").replace("||", " OR ")
    s = rename_vars(s)
    s = re.sub(r'\s+', ' ', s).strip()
    # drop one layer of fully-wrapping parens for readability
    while s.startswith("(") and _match_balanced(s, 0) == len(s):
        s = s[1:-1].strip()
    for _ in range(3):
        s = re.sub(r'\(\(([^()]+)\)\)', r'(\1)', s)     # collapse doubled parens
    s = s.replace("*", "·")
    s = re.sub(r'\s*·\s*', '·', s)                        # normalize middot spacing
    s = s.replace(" AND ", "  AND  ").replace(" OR ", "  OR  ")
    return s

def to_latex(readable):
    s = readable
    s = s.replace("~=", r"\approx ").replace(" AND ", r"\;\wedge\;").replace(" OR ", r"\;\vee\;")
    s = s.replace(" != ", r"\neq ").replace(" <= ", r"\leq ").replace(" >= ", r"\geq ")
    s = s.replace("·", r"\cdot ").replace("eps", r"\varepsilon")
    s = re.sub(r'\b(this|other|out|base|hi|lo)\.(re|im)([12])',
               r'\\mathrm{\1.\2}_{\3}', s)
    s = re.sub(r'\b(this|other|out)([12])\b', r'\\mathrm{\1}_{\2}', s)
    s = re.sub(r'\b([a-z]+)([12])\b', r'\1_{\2}', s)          # x1 -> x_{1}
    return s

def load():
    items = json.load(open(SRC))["labels"]
    rows = []
    for k, it in enumerate(items, 1):
        rows.append(dict(item_id="M%02d" % k, subject=it["subject"].replace("?", "."), mr_name=it["mr_name"],
                         author_label=it["author_label"], program=PROGRAMS.get(it["subject"].replace("?", "."), ""),
                         jir_raw=it["jir"], jor_raw=it["jor"],
                         jir=to_readable(it["jir"]), jor=to_readable(it["jor"])))
    return rows

def write_gold(rows):
    with open(HERE / "_gold_author_labels.csv", "w", newline="") as f:
        w = csv.DictWriter(f, ["item_id", "author_label"]); w.writeheader()
        for r in rows:
            w.writerow({"item_id": r["item_id"], "author_label": r["author_label"]})
    print("wrote _gold_author_labels.csv (hidden key)")

def write_sheet_csv(rows):
    with open(HERE / "rating_sheet_TEMPLATE.csv", "w", newline="") as f:
        w = csv.DictWriter(f, ["item_id", "subject", "program", "mr_name", "JIR", "JOR", "category", "notes"]); w.writeheader()
        for r in rows:
            w.writerow({"item_id": r["item_id"], "subject": r["subject"], "program": r["program"],
                        "mr_name": r["mr_name"], "JIR": r["jir"], "JOR": r["jor"], "category": "", "notes": ""})
    print("wrote rating_sheet_TEMPLATE.csv (CSV fallback; fill 'category' with a-j/orphan)")

def verify(rows):
    for r in rows[:8] + rows[20:24]:
        print("\n[%s] %s / %s" % (r["item_id"], r["subject"], r["mr_name"]))
        print("  JIR raw : %s" % r["jir_raw"][:150])
        print("  JIR read: %s" % r["jir"])
        print("  JOR read: %s" % r["jor"])

def write_csvs(rows):
    with open(HERE / "items_to_rate.csv", "w", newline="") as f:
        w = csv.DictWriter(f, ["item_id", "subject", "program", "mr_name", "JIR", "JOR"]); w.writeheader()
        for r in rows:
            w.writerow({"item_id": r["item_id"], "subject": r["subject"], "program": r["program"],
                        "mr_name": r["mr_name"], "JIR": r["jir"], "JOR": r["jor"]})
    with open(HERE / "items_raw.csv", "w", newline="") as f:
        w = csv.DictWriter(f, ["item_id", "JIR_raw", "JOR_raw"]); w.writeheader()
        for r in rows:
            w.writerow({"item_id": r["item_id"], "JIR_raw": r["jir_raw"], "JOR_raw": r["jor_raw"]})
    print("wrote items_to_rate.csv, items_raw.csv")

def write_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "rating"
    head = ["item_id", "subject", "program (what it computes)", "mr_name",
            "JIR (input relation)", "JOR (output relation)", "category", "notes"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["item_id"], r["subject"], r["program"], r["mr_name"], r["jir"], r["jor"], "", ""])
    dv = DataValidation(type="list", formula1='"a,b,c,d,e,f,g,h,i,j,orphan"', allow_blank=True,
                        showDropDown=False)
    dv.promptTitle = "MR family"; dv.prompt = "Pick ONE family a-j (or orphan). See CODEBOOK."
    dv.errorTitle = "Invalid"; dv.error = "Choose one of: a b c d e f g h i j orphan"
    ws.add_data_validation(dv)
    dv.add("G2:G%d" % (len(rows) + 1))
    widths = {"A": 7, "B": 22, "C": 40, "D": 18, "E": 44, "F": 38, "G": 11, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(HERE / "rating_sheet_TEMPLATE.xlsx")
    print("wrote rating_sheet_TEMPLATE.xlsx (category column F = dropdown a-j/orphan)")

TEX_HEAD = r"""\documentclass[10pt]{article}
\usepackage[a4paper,margin=1.8cm]{geometry}
\usepackage{amsmath,amssymb}
\usepackage[T1]{fontenc}
\setlength{\parindent}{0pt}
\sloppy
\begin{document}
\small
\section*{Items to rate (36 MRs) -- readable form}
Each MR runs the program twice. JIR = how the two \emph{inputs} relate; JOR = the
\emph{output} relation that must hold. Subscripts $1,2$ = first/second call;
\texttt{re/im} = real/imag part; \texttt{this/other} = operands; \texttt{out} =
return. Classify each into one MR family (a--j) or \texttt{orphan} (see CODEBOOK).
\bigskip

"""
TEX_FOOT = "\n\\end{document}\n"

def esc(s):
    s = s.replace("\\", "")
    for k, v in {"&": r"\&", "%": r"\%", "#": r"\#", "_": r"\_", "$": r"\$",
                 "^": r"\textasciicircum{}", "~": r"\textasciitilde{}", "|": r"\textbar{}",
                 "<": r"\textless{}", ">": r"\textgreater{}"}.items():
        s = s.replace(k, v)
    return s

def split_atoms(readable):
    """Emit each conjunct as its own $...$ so long formulas wrap (no overflow)."""
    parts = re.split(r'\s+(AND|OR)\s+', readable.strip())
    out = []
    for p in parts:
        if p == "AND":
            out.append(r"$\wedge$")
        elif p == "OR":
            out.append(r"$\vee$")
        elif p.strip():
            out.append("$" + to_latex(p.strip()) + "$")
    return " ".join(out) if out else r"$\,$"

def write_tex(rows):
    body = []
    for r in rows:
        body.append(r"\textbf{%s}\quad \texttt{%s / %s}\newline" %
                    (r["item_id"], esc(r["subject"]), esc(r["mr_name"])))
        body.append(r"\hspace*{1.5em}\emph{Program:} %s\newline" % esc(r["program"]))
        body.append(r"\hspace*{1.5em}JIR:\ %s\newline" % split_atoms(r["jir"]))
        body.append(r"\hspace*{1.5em}JOR:\ $\Rightarrow$\ %s\par\medskip" % split_atoms(r["jor"]))
    (HERE / "items_to_rate.tex").write_text(TEX_HEAD + "\n".join(body) + TEX_FOOT)
    print("wrote items_to_rate.tex")

def main():
    rows = load()
    if "--verify" in sys.argv:
        verify(rows); return
    write_csvs(rows); write_gold(rows); write_sheet_csv(rows); write_xlsx(rows); write_tex(rows)
    r = subprocess.run(["pdflatex", "-interaction=nonstopmode", "items_to_rate.tex"],
                       cwd=HERE, capture_output=True, text=True)
    print("pdflatex:", "OK" if (HERE / "items_to_rate.pdf").exists() else "FAILED")

if __name__ == "__main__":
    main()
