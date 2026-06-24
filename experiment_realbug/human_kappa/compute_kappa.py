#!/usr/bin/env python3
"""
Inter-rater kappa for the human MR-classification study -- TWO-LAYER scheme.

Raters assign each MR ONE MR family (Layer 2, ten families a-j) or `orphan`.
Each family rolls up to exactly one MetaPattern (Layer 1, five MetaPatterns), so
from a single family label we report agreement at BOTH layers:

  Layer 2 (families a-j + orphan):  Fleiss kappa + pairwise Cohen kappa
  Layer 1 (5 MetaPatterns + orphan): Fleiss kappa + pairwise Cohen kappa
                                     (families rolled up to MetaPatterns)
  Human vs author: at the MetaPattern layer (the author key is block-level, which
  rolls up cleanly to the 5 MetaPatterns; the LLM-panel audit in
  supplementary/S3_case_study can be re-scored at this layer for comparison).

Reads filled rater sheets from   human_kappa/ratings/rater_<name>.csv
(rows: item_id,category[,notes]; category = a family letter a-j, or 'orphan').

Pure standard library. Self-test: python3 compute_kappa.py --selftest
"""
import csv, glob, os, sys, random, itertools, pathlib

HERE = pathlib.Path(__file__).parent

# ---- Layer 2: the ten MR families (+ orphan) ----
FAMILIES = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j"]
FAM_CATS = FAMILIES + ["orphan"]
FAMILY_NAME = {
    "a": "equivariance", "b": "conservation", "c": "self-adjoint",
    "d": "adjoint-duality", "e": "time-reversal", "f": "static-order",
    "g": "dynamic-shape", "h": "convergence", "i": "accuracy-order",
    "j": "representation-invariance",
}
# ---- Layer 1: family -> MetaPattern ----
FAMILY_TO_MP = {"a": "G", "b": "G", "c": "T_star", "d": "T_star", "e": "T_rev",
                "f": "O_le", "g": "O_le", "h": "L_star", "i": "L_star", "j": "L_star"}
METAPATTERNS = ["G", "O_le", "T_star", "T_rev", "L_star"]
MP_CATS = METAPATTERNS + ["orphan"]
# ---- author key is block-level -> MetaPattern (refinements/extension fold in) ----
BLOCK_TO_MP = {"G": "G", "O_le": "O_le", "T_star": "T_star", "T_rev": "T_rev",
               "L_star": "L_star", "D_star": "O_le", "E_star": "L_star", "B_rel": "L_star"}

def roll(label_map, table):
    return {k: table.get(v, v) for k, v in label_map.items() if v}

# ---------- kappa primitives (generic over a category list) ----------
def cohen_kappa(a, b, cats):
    items = [i for i in (set(a) & set(b)) if a[i] and b[i]]
    n = len(items)
    if n == 0:
        return None, 0
    p_o = sum(1 for i in items if a[i] == b[i]) / n
    ca = {c: sum(1 for i in items if a[i] == c) / n for c in cats}
    cb = {c: sum(1 for i in items if b[i] == c) / n for c in cats}
    p_e = sum(ca[c] * cb[c] for c in cats)
    return (1.0 if p_e == 1 else (p_o - p_e) / (1 - p_e)), n

def fleiss_kappa(raters, cats, items=None):
    if items is None:
        items = [i for i in set.intersection(*[set(r) for r in raters]) if all(r.get(i) for r in raters)]
    n, N = len(raters), len(items)
    if N == 0 or n < 2:
        return None, N
    P, tot = [], {c: 0 for c in cats}
    for it in items:
        cnt = {c: 0 for c in cats}
        for r in raters:
            cnt[r[it]] += 1
        for c in cats:
            tot[c] += cnt[c]
        P.append((sum(v * v for v in cnt.values()) - n) / (n * (n - 1)))
    P_bar = sum(P) / N
    p = {c: tot[c] / (N * n) for c in cats}
    P_e = sum(v * v for v in p.values())
    return (1.0 if P_e == 1 else (P_bar - P_e) / (1 - P_e)), N

def fleiss_ci(raters, cats, B=2000, seed=0):
    items = [i for i in set.intersection(*[set(r) for r in raters]) if all(r.get(i) for r in raters)]
    if len(items) < 3:
        return None, None
    rng = random.Random(seed)
    ks = []
    for _ in range(B):
        k, _ = fleiss_kappa(raters, cats, items=[rng.choice(items) for _ in items])
        if k is not None:
            ks.append(k)
    ks.sort()
    return ks[int(0.025 * len(ks))], ks[int(0.975 * len(ks)) - 1]

def band(k):
    if k is None: return "n/a"
    return ("poor" if k < 0 else "slight" if k < .20 else "fair" if k < .40 else
            "moderate" if k < .60 else "substantial" if k < .80 else "almost perfect")

def majority(raters, items):
    out = {}
    for it in items:
        v = [r[it] for r in raters if r.get(it)]
        if v:
            out[it] = max(set(v), key=v.count)
    return out

# ---------- IO ----------
def load_sheet(path):
    """Read a rater sheet (.csv or .xlsx); return {item_id: family}."""
    pairs = []
    if str(path).endswith(".xlsx"):
        from openpyxl import load_workbook
        ws = load_workbook(path, data_only=True).active
        hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        ii = hdr.index("item_id") if "item_id" in hdr else 0
        ci = hdr.index("category") if "category" in hdr else len(hdr) - 1
        for r in ws.iter_rows(min_row=2):
            iid = r[ii].value if ii < len(r) else None
            cat = r[ci].value if ci < len(r) else None
            pairs.append((str(iid).strip() if iid is not None else "",
                          str(cat).strip() if cat is not None else ""))
    else:
        for row in csv.DictReader(open(path, newline="")):
            pairs.append(((row.get("item_id") or "").strip(), (row.get("category") or "").strip()))
    d = {}
    for iid, cat in pairs:
        if iid and cat:
            if cat not in FAM_CATS:
                print("  WARNING %s: %s has unknown family %r (expected a-j or orphan; ignored)"
                      % (os.path.basename(str(path)), iid, cat)); continue
            d[iid] = cat
    return d

def load_gold():
    p = HERE / "_gold_author_labels.csv"
    return {r["item_id"]: r["author_label"] for r in csv.DictReader(open(p))} if p.exists() else None

def kappa_block(title, raters, names, cats):
    print("\n== %s ==" % title)
    for i, j in itertools.combinations(range(len(raters)), 2):
        k, n = cohen_kappa(raters[i], raters[j], cats)
        print("  %-10s vs %-10s : kappa=%s (n=%d, %s)" %
              (names[i], names[j], "%.3f" % k if k is not None else "n/a", n, band(k)))
    k, N = fleiss_kappa(raters, cats)
    lo, hi = fleiss_ci(raters, cats) if k is not None else (None, None)
    ci = "" if lo is None else "  95%% CI [%.3f, %.3f]" % (lo, hi)
    print("  Fleiss kappa=%s (n=%d, r=%d, c=%d, %s)%s" %
          ("%.3f" % k if k is not None else "n/a", N, len(raters), len(cats), band(k), ci))

def report(files):
    raters, names = [], []
    for fp in files:
        d = load_sheet(fp)
        if d:
            raters.append(d); names.append(pathlib.Path(fp).stem.replace("rater_", ""))
    if len(raters) < 2:
        print("Need >=2 rater sheets with filled families. Found %d." % len(raters)); return
    common = [i for i in set.intersection(*[set(r) for r in raters]) if all(r.get(i) for r in raters)]
    print("Raters (%d): %s\nItems rated by all: %d" % (len(raters), ", ".join(names), len(common)))

    kappa_block("Layer 2 -- MR family agreement (a-j + orphan)", raters, names, FAM_CATS)

    mp_raters = [roll(r, FAMILY_TO_MP) for r in raters]
    kappa_block("Layer 1 -- MetaPattern agreement (5 MetaPatterns + orphan, rolled up)",
                mp_raters, names, MP_CATS)

    gold = load_gold()
    if gold:
        gold_mp = roll(gold, BLOCK_TO_MP)
        maj_mp = majority(mp_raters, common)
        k, n = cohen_kappa(maj_mp, gold_mp, MP_CATS)
        print("\n== Human vs author (MetaPattern layer) ==")
        print("  human-majority vs author : kappa=%s (n=%d, %s)" %
              ("%.3f" % k if k is not None else "n/a", n, band(k)))
        for nm, r in zip(names, mp_raters):
            k, n = cohen_kappa(r, gold_mp, MP_CATS)
            print("  %-10s vs author : kappa=%s (n=%d, %s)" %
                  (nm, "%.3f" % k if k is not None else "n/a", n, band(k)))

    print("\n== Family-level disagreements ==")
    found = False
    for it in sorted(common):
        labs = {nm: r[it] for nm, r in zip(names, raters)}
        if len(set(labs.values())) > 1:
            found = True
            mp = {nm: FAMILY_TO_MP.get(v, v) for nm, v in labs.items()}
            same_mp = "(same MetaPattern)" if len(set(mp.values())) == 1 else "(DIFFERENT MetaPattern)"
            extra = "  [author block=%s -> MP=%s]" % (gold.get(it, "?"), BLOCK_TO_MP.get(gold.get(it, ""), "?")) if gold else ""
            print("  %s: %s %s%s" % (it, labs, same_mp, extra))
    if not found:
        print("  none")

def selftest():
    print("=" * 70)
    print("SYNTHETIC SELF-TEST -- fabricated raters, NOT human data. Proves the")
    print("pipeline runs; the kappa below is meaningless as evidence.")
    print("=" * 70)
    gold = load_gold()
    if not gold:
        print("(_gold_author_labels.csv missing; run make_items.py first.)"); return
    # plausible family per item from the block key (block -> a representative family)
    block_fam = {"G": "a", "O_le": "f", "T_star": "c", "T_rev": "e",
                 "L_star": "h", "D_star": "g", "E_star": "i", "B_rel": "j"}
    rng = random.Random(42)
    d = HERE / "ratings_selftest"; d.mkdir(exist_ok=True)
    files = []
    for nm, noise in [("synthA", .12), ("synthB", .18), ("synthC", .22)]:
        fp = d / ("rater_%s.csv" % nm)
        with open(fp, "w", newline="") as fh:
            w = csv.writer(fh); w.writerow(["item_id", "category", "notes"])
            for iid, blk in gold.items():
                base = block_fam.get(blk, "a")
                lab = rng.choice([c for c in FAMILIES if c != base]) if rng.random() < noise else base
                w.writerow([iid, lab, "SYNTHETIC"])
        files.append(str(fp))
    report(files)
    print("\n(Delete experiment_realbug/human_kappa/ratings_selftest/ before real use.)")

if __name__ == "__main__":
    if "--selftest" in sys.argv:
        selftest()
    else:
        files = sorted(glob.glob(str(HERE / "ratings" / "rater_*.csv")) +
                       glob.glob(str(HERE / "ratings" / "rater_*.xlsx")))
        if not files:
            print("No rater sheets in ./ratings/. Each rater copies rating_sheet_TEMPLATE.csv")
            print("to ratings/rater_<name>.csv, fills 'category' with ONE family letter")
            print("(a,b,c,d,e,f,g,h,i,j) or 'orphan', then run this script.")
            print("\nPipeline demo on fabricated data: python3 compute_kappa.py --selftest")
        else:
            report(files)
